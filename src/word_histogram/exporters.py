"""
Export functionality for word histogram data.

This module provides functions to export word frequency data
to various file formats including text and Excel.
"""
import openpyxl
from pathlib import Path
from typing import List, Tuple, Union


def save_to_text(word_counts: List[Tuple[str, int]], output_file: Union[str, Path]) -> None:
    """
    Save word counts to a text file.
    
    Args:
        word_counts: List of (word, count) tuples
        output_file: Path to the output text file
        
    Returns:
        None
    """
    with open(output_file, 'w', encoding='utf-8') as file:
        for word, count in word_counts:
            file.write(f"{count:<10} {word}\n")


def save_to_excel(word_counts: List[Tuple[str, int]], output_file: Union[str, Path]) -> None:
    """
    Save word counts to an Excel file.
    
    Args:
        word_counts: List of (word, count) tuples
        output_file: Path to the output Excel file
        
    Returns:
        None
        
    Raises:
        ImportError: If openpyxl is not installed
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Add headers
    ws.cell(row=1, column=1, value="Word")
    ws.cell(row=1, column=2, value="Count")
    
    # Add data
    for row, (word, count) in enumerate(word_counts, start=2):
        ws.cell(row=row, column=1, value=word)
        ws.cell(row=row, column=2, value=count)
    
    wb.save(output_file)